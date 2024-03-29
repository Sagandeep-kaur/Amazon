from openpyxl import Workbook, load_workbook
import re
import inspect
import logging
import softest
class Utils():
    #softest.TestCase
      def __init__(self,browser):
         self.browser= browser

      def assert_list_items(self,list, value):

          for stop in list:
              assert (stop.text) == value
              print("assert passed")

      def scroll_to_elemInList_AndClick(self,list):

          for elem in list:

              self.browser.execute_script("arguments[0].scrollIntoView();", elem)
              self.browser.execute_script("arguments[0].click();", elem)




      def read_data_from_excel(file_name,sheet):

          datalist = []
          wb = load_workbook(filename=file_name)
          sh = wb[sheet]
          row_ct = sh.max_row
          col_ct = sh.max_column

          for i in range(2, row_ct + 1):
              row = []
              for j in range(1, col_ct + 1):
                  row.append(sh.cell(row=i, column=j).value)
              datalist.append(row)

          return datalist

      def getintegerwordfromstring(self,string):

          global elem
          list = re.findall("\d+", string)
          #print(list)
          for i in list:
              elem = int(i)
          return elem

      def custom_logger(logLevel=logging.DEBUG):
        #Set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log", mode='a')
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger